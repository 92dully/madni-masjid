"""
Build prayer-times.js from:
  - GCM Prayer Times.xlsx  -> daily *begin* times (Hanafi Asr, column "Asr Begins (II)")
  - The 2026 jamaat image  -> *jamaat* times (encoded in the rule tables below)

Maghrib jamaat = Maghrib begin time.

Run:
    python _build_data.py

Output: `prayer-times.js`, which exposes:
    window.PRAYER_TIMES_DATA = [ { date, day, fajr_begin, fajr_jamaat, ... }, ... ]

Special override (confirmed with masjid):
  Until 8th August 2026, Fajr Jamaat = 01:35
  (The image shows 01:50 from 04-May, but the masjid has confirmed 01:35 until 09-Aug.)
"""
import json
import openpyxl
from datetime import time, datetime


# ---------------------------------------------------------------------------
# FAJR JAMAAT RULES — transcribed from 2026 timetable image
# Each entry: ("MM-DD", "HH:MM") means "from this date, jamaat is at this time"
# None = Refer to Ramadan timetable
# ---------------------------------------------------------------------------

FAJR_RULES = [
    ("01-01", None),        # Ramadan period at start of year
    ("03-22", "05:30"),
    ("03-29", "06:00"),
    ("04-07", "05:45"),
    ("04-12", "05:30"),
    ("04-18", "05:15"),
    ("04-23", "05:00"),
    ("05-01", "04:45"),
    ("05-04", "01:35"),     # Image shows 01:50 but masjid confirmed 01:35 until 09-Aug
    ("08-09", "05:00"),
    ("08-16", "05:15"),
    ("08-23", "05:30"),
    ("08-31", "05:45"),
    ("09-08", "06:00"),
    ("09-16", "06:15"),
    ("09-23", "06:30"),
    ("10-01", "06:45"),
    ("10-08", "07:00"),
    ("10-25", "06:45"),
    ("11-06", "07:00"),
]

# ---------------------------------------------------------------------------
# ZOHAR (ZOHAR) JAMAAT RULES — 2026
# Zohar column covers Jumu'ah too.
# 01 Jan – 28 Mar: 1pm (GMT)
# 29 Mar – 24 Oct: 2pm (BST)
# 25 Oct – 31 Dec: 1pm (GMT)
# ---------------------------------------------------------------------------

ZOHAR_RULES = [
    ("01-01", "13:00"),     # GMT: 1pm
    ("03-29", "14:00"),     # BST starts 29 Mar 2026
    ("10-25", "13:00"),     # BST ends   25 Oct 2026
]

# ---------------------------------------------------------------------------
# ASR JAMAAT RULES — transcribed from 2026 timetable image
# ---------------------------------------------------------------------------

ASR_RULES = [
    ("01-19", "15:15"),
    ("02-03", "15:45"),
    ("02-16", None),        # Ramadan period
    ("03-21", "17:30"),
    ("03-29", "19:00"),
    ("05-23", "19:15"),
    ("06-27", "20:00"),
    ("08-10", "19:00"),
    ("09-03", "18:00"),
    ("09-11", "17:45"),
    ("10-01", "17:00"),
    ("10-08", "16:45"),
    ("10-25", "15:15"),
    ("11-16", "14:45"),
]

# ---------------------------------------------------------------------------
# ISHA JAMAAT RULES — transcribed from 2026 timetable image
# ---------------------------------------------------------------------------

ISHA_RULES = [
    ("01-29", "19:15"),
    ("02-07", "19:30"),
    ("02-15", "19:45"),
    ("02-20", None),        # Ramadan period
    ("03-20", "20:45"),
    ("03-29", "22:00"),
    ("04-22", "22:15"),
    ("05-02", "22:45"),
    ("05-13", "23:00"),
    ("05-25", "23:15"),
    ("07-14", "23:00"),
    ("07-25", "22:45"),
    ("08-04", "22:30"),
    ("08-13", "22:15"),
    ("08-29", "22:00"),
    ("09-19", "21:45"),
    ("09-24", "21:30"),
    ("09-29", "21:15"),
    ("10-04", "21:00"),
    ("10-09", "20:45"),
    ("10-15", "20:30"),
    ("10-21", "20:15"),
    ("10-25", "19:00"),
]


_NOT_FOUND = object()


def jamaat_for(rules, dt):
    """Return the jamaat time string for a date by walking the MM-DD rules."""
    md = dt.strftime("%m-%d")
    chosen = _NOT_FOUND
    for rule_md, value in rules:
        if rule_md <= md:
            chosen = value
        else:
            break
    if chosen is _NOT_FOUND:
        # Before the year's first rule: use the last rule of the prior year
        chosen = rules[-1][1] if rules else None
    return chosen  # may be None for Ramadan placeholders


def fmt(t):
    if t is None:
        return None
    if isinstance(t, (time, datetime)):
        return t.strftime("%H:%M")
    return str(t)


def main():
    wb = openpyxl.load_workbook("GCM_Prayer_Times.xlsx", data_only=True)
    # Try both sheet name variants
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    days = []
    for row in range(3, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if date_val is None:
            continue

        maghrib_begin = fmt(ws.cell(row=row, column=11).value)

        day = {
            "date":           date_val.strftime("%Y-%m-%d"),
            "day":            ws.cell(row=row, column=2).value,
            "fajr_begin":     fmt(ws.cell(row=row, column=3).value),
            "fajr_jamaat":    jamaat_for(FAJR_RULES, date_val),
            "sunrise":        fmt(ws.cell(row=row, column=5).value),
            "zohar_begin":    fmt(ws.cell(row=row, column=6).value),
            "zohar_jamaat":   jamaat_for(ZOHAR_RULES, date_val),
            # Hanafi Asr (column 9 = "Asr Begins (II)")
            "asr_begin":      fmt(ws.cell(row=row, column=9).value),
            "asr_jamaat":     jamaat_for(ASR_RULES, date_val),
            "maghrib_begin":  maghrib_begin,
            "maghrib_jamaat": maghrib_begin,
            "isha_begin":     fmt(ws.cell(row=row, column=13).value),
            "isha_jamaat":    jamaat_for(ISHA_RULES, date_val),
        }
        days.append(day)

    payload = "window.PRAYER_TIMES_DATA = " + json.dumps(days, indent=2) + ";\n"
    with open("prayer-times.js", "w", encoding="utf-8") as f:
        f.write(payload)

    print(
        f"Wrote prayer-times.js with {len(days)} days "
        f"({days[0]['date']} to {days[-1]['date']})."
    )


if __name__ == "__main__":
    main()
